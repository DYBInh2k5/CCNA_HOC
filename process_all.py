import sys
import os
import pypdf
import docx
import openpyxl
import io

# Set stdout to UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def extract_pdf(file_path):
    try:
        reader = pypdf.PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += (page.extract_text() or "") + "\n"
        return text
    except Exception as e:
        return f"Error reading PDF: {e}"

def extract_docx(file_path):
    try:
        doc = docx.Document(file_path)
        text = "\n".join([para.text for para in doc.paragraphs])
        return text
    except Exception as e:
        return f"Error reading DOCX: {e}"

def extract_xlsx(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            text += f"--- Sheet: {sheet} ---\n"
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
        return text
    except Exception as e:
        return f"Error reading XLSX: {e}"

def extract_txt(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    except Exception as e:
        return f"Error reading TXT: {e}"

def process_directory(root_dir):
    summary = []
    for root, dirs, files in os.walk(root_dir):
        for file in files:
            file_path = os.path.join(root, file)
            ext = os.path.splitext(file)[1].lower()
            
            # Skip the script itself and non-target files
            if file in ['extract_text.py', 'process_all.py'] or ext == '.exe':
                continue
                
            print(f"\n{'='*20}\nFILE: {file_path}\n{'='*20}")
            
            content = ""
            if ext == '.pdf':
                content = extract_pdf(file_path)
            elif ext == '.docx' or ext == '.doc':
                content = extract_docx(file_path)
            elif ext == '.xlsx':
                content = extract_xlsx(file_path)
            elif ext == '.txt':
                content = extract_txt(file_path)
            else:
                content = f"Skipping unsupported format: {ext}"
            
            # Print first 2000 chars of each file to avoid hitting terminal limits but still get enough context
            print(content[:2000] + ("..." if len(content) > 2000 else ""))

if __name__ == "__main__":
    target_dir = "Quantri Mang"
    if not os.path.exists(target_dir):
        print(f"Directory not found: {target_dir}")
        sys.exit(1)
    
    process_directory(target_dir)
