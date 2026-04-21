import sys
import os
import pypdf
import docx
import openpyxl

def extract_pdf(file_path):
    try:
        reader = pypdf.PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        return f"Error reading PDF: {e}"

def extract_docx(file_path):
    try:
        doc = docx.Document(file_path)
        text = "\n".join([para.text for para in doc.paragraphs])
        return text
    except Exception as e:
        return f"Error reading DOCX: {e}"

def extract_xlsx(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            text += f"--- Sheet: {sheet} ---\n"
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
        return text
    except Exception as e:
        return f"Error reading XLSX: {e}"

if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    if len(sys.argv) < 2:
        print("Usage: python extract_text.py <file_path>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        sys.exit(1)
    
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".pdf":
        print(extract_pdf(file_path))
    elif ext == ".docx":
        print(extract_docx(file_path))
    elif ext == ".xlsx":
        print(extract_xlsx(file_path))
    else:
        print(f"Unsupported file format: {ext}")
